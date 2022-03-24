#!/usr/bin/env python3

import csv
import json
import sys
import xml.etree.ElementTree as ET

from dataclasses import dataclass
from datetime import datetime
from statistics import mean

from openpyxl import Workbook

if sys.version_info.major < 3 or (sys.version_info.major == 3 and sys.version_info.minor < 8):
    print('This script requires the walrus operator to be supported.', file=sys.stderr)
    print('Please use Python 3.8 or greater.', file=sys.stderr)
    sys.exit(1)


@dataclass
class Sample:
    name: str
    barcode_sequence: str
    pf_clusters: list[int]
    percent_of_lanes: list[float]
    yield_in_mbases: list[int]
    percent_geq_q30_bases: list[float]
    mean_quality_score: list[float]


def performance_excel_workbook(stats_json, coverage_summary, run_info, protocol):
    stats_object = json.load(stats_json)
    samples = dict()
    for lane in stats_object["ConversionResults"]:
        total_pf_clusters = lane["TotalClustersPF"]
        for sample_dict in lane["DemuxResults"]:
            read_metrics_yield = sum([x["Yield"] for x in sample_dict["ReadMetrics"]])
            read_metrics_yield_q30 = sum([x["YieldQ30"] for x in sample_dict["ReadMetrics"]])
            read_metrics_quality_score_sum = sum([x["QualityScoreSum"] for x in sample_dict["ReadMetrics"]])
            try:
                percent_geq_q30_bases = read_metrics_yield_q30 * 100 / read_metrics_yield
            except ZeroDivisionError:
                percent_geq_q30_bases = 0
            try:
                mean_quality_score = read_metrics_quality_score_sum / read_metrics_yield
            except ZeroDivisionError:
                mean_quality_score = 0
            if (sample_id := sample_dict["SampleId"]) in samples:
                if sample_dict["SampleName"] == samples[sample_id].name and sample_dict["IndexMetrics"][0]["IndexSequence"] == samples[sample_id].barcode_sequence:
                    samples[sample_id].pf_clusters.append(sample_dict["NumberReads"])
                    samples[sample_id].percent_of_lanes.append(sample_dict["NumberReads"]*100/total_pf_clusters)
                    samples[sample_id].yield_in_mbases.append(sample_dict["Yield"]//1000000)
                    samples[sample_id].percent_geq_q30_bases.append(percent_geq_q30_bases)
                    samples[sample_id].mean_quality_score.append(mean_quality_score)
                else:
                    raise ValueError(f"mismatch with {sample_id}")
            else:
                samples[sample_id] = Sample(
                        name = sample_dict["SampleName"],
                        barcode_sequence = sample_dict["IndexMetrics"][0]["IndexSequence"],
                        pf_clusters = [sample_dict["NumberReads"]],
                        percent_of_lanes = [sample_dict["NumberReads"]*100/total_pf_clusters],
                        yield_in_mbases = [sample_dict["Yield"]//1000000],
                        percent_geq_q30_bases = [percent_geq_q30_bases],
                        mean_quality_score = [mean_quality_score]
                        )

    sample_names = sorted(list(samples))

    coverage_dialect = csv.Sniffer().sniff(coverage_summary.readline())
    coverage_summary.seek(0)
    coverage_reader = csv.DictReader(coverage_summary, dialect=coverage_dialect)
    coverage_samples = [row for row in coverage_reader]
    coverage_lookup = {j[coverage_reader.fieldnames[0]]:i for i, j in enumerate(coverage_samples)}

    run_info_tree = ET.parse(run_info)
    run_info_root = run_info_tree.getroot()
    sequence_date_raw = run_info_root[0].find("Date").text
    try:
        sequence_date = datetime.strptime(sequence_date_raw, "%y%m%d").strftime("%Y-%m-%d")
    except ValueError:
        sequence_date = datetime.strptime(sequence_date_raw.split(" ")[0], "%m/%d/%Y").strftime("%Y-%m-%d")
    instrument_name = run_info_root[0].find("Instrument").text

    worksheet_header = ["Sample"] + coverage_reader.fieldnames[1:] + ["Barcode Sequences", "PF Clusters", "% of the lanes", "Yield (Mbases)", "% >= Q30 Bases", "Mean Quality Score", "Date of Sequence", "Instrument", "Protocol"]
    wb = Workbook()
    ws = wb.active
    for i, j in enumerate(worksheet_header):
        ws.cell(column=i+1, row=1, value=j)
    
    for idx, sample_name in enumerate(sample_names):
        row_template = [sample_name]
        try:
            full_sample_name = [key for key in coverage_lookup if key.startswith(sample_name)][0]
        except IndexError:
            for fieldname in coverage_reader.fieldnames[1:]:
                row_template.append(0)
        else:
            coverage_sample = coverage_samples[coverage_lookup[full_sample_name]]
            for fieldname in coverage_reader.fieldnames[1:]:
                row_template.append(coverage_sample[fieldname])

        row_template.append(samples[sample_name].barcode_sequence)
        row_template.append(sum(samples[sample_name].pf_clusters))
        row_template.append(mean(samples[sample_name].percent_of_lanes))
        row_template.append(sum(samples[sample_name].yield_in_mbases))
        row_template.append(mean(samples[sample_name].percent_geq_q30_bases))
        row_template.append(mean(samples[sample_name].mean_quality_score))

        row_template.append(sequence_date)
        row_template.append(instrument_name)
        row_template.append(protocol)
        
        for i,j in enumerate(row_template):
            ws.cell(column=i+1, row=idx+2, value=j)

    return wb

def lineage_excel_workbook(pangolin_summary, nextclade_summary, vadr_annotations):
    wb = Workbook()
    ws_pangolin = wb.active
    ws_pangolin.title = "Pangolin"

    pangolin_dialect = csv.Sniffer().sniff(pangolin_summary.readline())
    pangolin_summary.seek(0)
    for row in filter(None, csv.reader(pangolin_summary, dialect=pangolin_dialect)):
        ws_pangolin.append(row)

    ws_nextclade = wb.create_sheet("Nextclade")
    nextclade_dialect = csv.Sniffer().sniff(nextclade_summary.readline())
    nextclade_summary.seek(0)
    for row in filter(None, csv.reader(nextclade_summary, dialect=nextclade_dialect)):
        ws_nextclade.append(row)

    ws_vadr_annotations = wb.create_sheet("VADR Annotations")
    ws_vadr_annotations.append([
        "Name",
        "Length",
        "Pass/Fail",
        "Annotated?",
        "Best Model",
        "Group",
        "Subgroup",
        "# of Features Annotated",
        "# of Features Not Annotated",
        "# of 5' Truncated Features",
        "# of 3' Truncated Features",
        "# of Per-Feature Alerts",
        "Per-Sequence Alerts"
        ])
    for row in filter(lambda x: not x.startswith("#"), vadr_annotations.readlines()):
        ws_vadr_annotations.append(row.strip().split()[1:])

    return wb


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description = "Make some excel files related to performance and lineage.")
    parser.add_argument("prefix", type=str, help="Prefix for output excel files.")

    subparsers = parser.add_subparsers(help="Performance or lineage", dest="which")

    performance_parser = subparsers.add_parser("performance", help="Generate performance excel")
    performance_parser.add_argument("--stats_json", type=argparse.FileType("r"), required=True, help="Stats.json produced by bcl2fastq")
    performance_parser.add_argument("--coverage_summary", type=argparse.FileType("r"), required=True, help="TXT file produced by summarizecoverage.sh")
    performance_parser.add_argument("--run_info", type=argparse.FileType("r"), required=True, help="RunInfo.xml produced by the Illumina machine")
    performance_parser.add_argument("--protocol", type=str, required=True, help="This run's protocol")

    lineage_parser = subparsers.add_parser("lineage", help="Generate lineage excel")
    lineage_parser.add_argument("--pangolin_summary", type=argparse.FileType("r"), required=True, help="Tabular summary of Pangolin lineages")
    lineage_parser.add_argument("--nextclade_summary", type=argparse.FileType("r"), required=True, help="Tabular summary of Nextclade lineages")
    lineage_parser.add_argument("--vadr_annotations", type=argparse.FileType("r"), required=True, help="Tabular summary of VADR annotations")

    args = parser.parse_args()

    if args.which == "performance":
        performance_excel_workbook(args.stats_json, args.coverage_summary, args.run_info, args.protocol).save(f"{args.prefix}_performance.xlsx")
    if args.which == "lineage":
        lineage_excel_workbook(args.pangolin_summary, args.nextclade_summary, args.vadr_annotations).save(f"{args.prefix}_lineage.xlsx")
